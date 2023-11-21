#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active
wr = ws.max_row
score = []
sorted_score = []
rank = [0, 0]
countF = 0

for r in range(2, wr + 1):
    midterm = ws.cell(row = r, column = 3).value * 0.3
    final = ws.cell(row = r, column = 4).value * 0.35
    homework = ws.cell(row = r, column = 5).value * 0.34
    attendance = ws.cell(row = r, column = 6).value * 1
    total = midterm + final + homework + attendance
    ws.cell(row=r, column=7, value=total)
    score.append(total)
    if total < 40:
        countF += 1
    r += 1

sorted_score = sorted(score, reverse = True)
for s in score:
    rank.append(sorted_score.index(s) + 1)

countA = int((wr - 1) * 0.3)
countB = int((wr - 1) * 0.7) - countA
countC = (wr - 1) - countA - countB - countF
#print(countA,countB,countC,countF)
for r in range(2, wr + 1):
    if ws.cell(row=r, column=7).value < 40:
        ws.cell(row=r, column=8, value='F')
    elif rank[r] <= countA / 2:
        ws.cell(row=r, column=8, value='A+')
    elif rank[r] <= countA:
        ws.cell(row=r, column=8, value='A0')
    elif rank[r] <= countA + countB / 2:
        ws.cell(row=r, column=8, value='B+')
    elif rank[r] <= countA + countB:
        ws.cell(row=r, column=8, value='B0')
    elif rank[r] <= countA + countB + countC / 2:
        ws.cell(row=r, column=8, value='C+')
    else:
        ws.cell(row=r, column=8, value='C0')

wb.save(filename = 'student.xlsx')
wb.close()