from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active
wr = ws.max_row

print(wr)

for r in range(2, wr + 1):
    midterm = ws.cell(row = r, column = 3).value * 30 / 100
    final = ws.cell(row = r, column = 4).value * 35 / 100
    homework = ws.cell(row = r, column = 5).value * 34 / 100
    attendance = ws.cell(row = r, column = 6).value * 1
    total = midterm + final + homework + attendance
    ws.cell(row = r, column = 7, value = total)
    r += 1

wb.save(filename = 'student.xlsx')
wb.close()